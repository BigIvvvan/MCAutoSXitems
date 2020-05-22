import openpyxl  # 导入模块

wb = openpyxl.load_workbook('item.xlsx')  # 打开表格


def readxl(county, sheet):
    ilist = []  # 创建返回属性列表
    if sheet == 'item1':  # 读取武器列表
        sh = wb['武器']
        countx = 1
        Maxc = sh.max_column
        while countx <= Maxc:
            read_data = sh.cell(row=county, column=countx)
            ilist.append(str(read_data.value))
            countx = countx + 1
        return ilist
    elif sheet == 'item2':  # 读取装备列表
        sh = wb['装备']
        countx = 1
        Maxc = sh.max_column
        while countx <= Maxc:
            read_data = sh.cell(row=county, column=countx)
            ilist.append(str(read_data.value))
            countx = countx + 1
        return ilist
    else:
        print('error')


def write_item1():
    txt_item1 = ''
    sh = wb['武器']
    Maxc = sh.max_column
    Maxr = sh.max_row
    county = 1
    countx = 1
    lore_list =  readxl(1,'item1')# 读取属性列表

    print(lore_list)
    county = county + 1
    while county <= Maxr:
        txt1 = ''  # 创建临时txt，用来寄存单个武器
        ilist = readxl(county, 'item1')
        UID = ilist[0]
        name = ilist[1]
        aa =str(name) + ':\n  Name:' + str(name) + '\n  ID:' + str(UID) + '\n  Lore:\n'
        txt1=txt1+aa
        a = 1
        for i in ilist:

            if 'None' in i:
                a = a + 1
                continue
            elif '_' in i:
                i.insert(1, '<r:')
                if '%' in i:
                    i.insert(-2, '>')
                else:
                    i.insert(-1, '>')
            else:
                i=i
            lore = '  - \''
            lore=lore+lore_list[a]
            lore=lore+': ' + i
            lore=lore+'\'\n'
            txt1=txt1+lore
            a = a + 1
        txt1=txt1+lore + '  - &e获得时间：<t:0>\n  ItemFlagList:\n  - HIDE_PLACED_ON\n  Unbreakable: true\n\n'
        txt_item1= txt_item1+txt1

        county = county + 1

    return (txt_item1)
def write_item2():
    txt_item1 = ''
    sh = wb['装备']
    Maxc = sh.max_column
    Maxr = sh.max_row
    county = 1
    countx = 1
    lore_list =  readxl(1,'item2')# 读取属性列表

    print(lore_list)
    county = county + 1
    while county <= Maxr:
        txt1 = ''  # 创建临时txt，用来寄存单个武器
        ilist = readxl(county, 'item2')
        UID = ilist[0]
        name = ilist[1]
        aa =str(name) + ':\n  Name:' + str(name) + '\n  ID:' + str(UID) + '\n  Lore:\n'
        txt1=txt1+aa
        a = 1
        for i in ilist:

            if 'None' in i:
                a = a + 1
                continue
            elif '_' in i:
                i.insert(1, '<r:')
                if '%' in i:
                    i.insert(-2, '>')
                else:
                    i.insert(-1, '>')
            else:
                i=i
            lore = '  - \''
            lore=lore+lore_list[a]
            lore=lore+': ' + i
            lore=lore+'\'\n'
            txt1=txt1+lore
            a = a + 1
        txt1=txt1+lore + '  - &e获得时间：<t:0>\n  ItemFlagList:\n  - HIDE_PLACED_ON\n  Unbreakable: true'
        print(txt1)
        county = county + 1
        txt_item1 = txt_item1+ txt1

    return (txt_item1)

with open('RPGitem.txt','w') as f:
    f.write(write_item1())
    f.write(write_item2())



